#!/usr/bin/env python3
"""Convert the authoritative Style Tracker workbook tabs into importable TSV.

The output matches public.style_tracker_rows. Database replacement is deliberately
left to the caller so it can load a staging table and swap both tabs atomically.
"""

from __future__ import annotations

import argparse
import csv
import json
from datetime import date, datetime, time
from decimal import Decimal
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


WORKBOOK_ID = "1ZL6cEwydC0cWSGP2I92uILn1ixILr_qAeDfDfD6F214"

SHEETS = {
    "License.Style": {
        "tracker_type": "licensed",
        "business_columns": ("B", "C", "D", "E", "F", "G", "K", "L", "M"),
        "typed": {
            "sku": "B", "group_id": "C", "description": "D", "customer": "E",
            "designer": "F", "commissioned": "G", "upc": "K", "customer_sku": "L",
            "licensor": "M", "license_status": "N", "concept_status": "O",
            "pre_production_status": "X", "production_status": "AC",
            "default_vendor": "AD", "discontinued": "AK", "notes": "AU",
        },
        "legacy": {
            "A": "print_fair_row", "B": "style_sku", "C": "group_id", "D": "description",
            "E": "originally_designed_for", "F": "designer", "G": "commissioned",
            "H": "rfq_code", "I": "legacy_ba", "J": "ba", "K": "upc", "L": "customer_sku",
            "M": "licensor", "N": "license_status", "O": "concept_sent",
            "P": "concept_resubmit", "Q": "concept_resubmitted", "R": "concept_approval",
            "S": "concept_approved_with_comments", "T": "request_pre_production_sample",
            "U": "sample_vendor", "W": "sample_photos_received", "X": "pre_production_sent",
            "Y": "pre_production_resubmit", "Z": "pre_production_resubmitted",
            "AA": "pre_production_approved_comment", "AB": "pre_production_approval",
            "AC": "production_approval", "AD": "default_vendor_sales", "AE": "ordered_cont_sample",
            "AF": "ordered_proff_photos", "AG": "ordered_test_report", "AH": "professional_photos",
            "AI": "test_report", "AJ": "catalog_image", "AK": "discontinued",
            "AL": "customer_exclusive", "AM": "annual_samples_need_order",
            "AN": "annual_samples_ordered", "AO": "contractual_samples_reorder",
            "AQ": "tp_assigned", "AU": "note",
        },
    },
    "Generic.Style": {
        "tracker_type": "generic",
        "business_columns": ("B", "C", "D", "E", "F", "G", "H", "I", "J"),
        "typed": {
            "sku": "B", "group_id": "C", "description": "D", "customer": "E",
            "designer": "F", "commissioned": "G", "upc": "H", "customer_sku": "I",
            "licensor": "J", "royalty": "M", "concept_status": "N",
            "pre_production_status": "V", "production_status": "X",
            "default_vendor": "Y", "discontinued": "AF", "notes": None,
        },
        "legacy": {
            "B": "style_sku", "C": "group_id", "D": "description", "E": "special_customer",
            "F": "designer", "G": "commissioned", "H": "upc", "I": "customer_sku",
            "J": "licensor", "M": "royalty", "N": "concept_sent", "O": "concept_resubmit",
            "P": "concept_resubmitted", "Q": "concept_approval", "R": "request_pre_production_sample",
            "S": "sample_vendor", "T": "rfq_code", "U": "sample_received",
            "V": "pre_production_sent", "W": "pre_production_approval",
            "X": "production_approval", "Y": "default_vendor_sales", "Z": "ordered_david_sample",
            "AA": "ordered_proff_photos", "AB": "ordered_test_report", "AC": "professional_photos",
            "AD": "test_report", "AE": "catalog_image", "AF": "discontinued",
        },
    },
}

TYPED_FIELDS = (
    "sku", "group_id", "description", "customer", "designer", "commissioned", "upc",
    "customer_sku", "licensor", "license_status", "royalty", "concept_status",
    "pre_production_status", "production_status", "default_vendor", "discontinued", "notes",
)


def display_value(value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, bool):
        return value
    if isinstance(value, datetime):
        if value.time() == time.min:
            return f"{value.month}/{value.day}/{value.year}"
        return value.isoformat(sep=" ")
    if isinstance(value, date):
        return f"{value.month}/{value.day}/{value.year}"
    if isinstance(value, Decimal):
        return format(value, "f").rstrip("0").rstrip(".")
    if isinstance(value, float):
        if value.is_integer():
            return str(int(value))
        return format(value, ".15g")
    if isinstance(value, int):
        return str(value)
    text = str(value).strip()
    return text or None


def is_business_value(value: Any) -> bool:
    value = display_value(value)
    if value is None or value is False:
        return False
    return str(value).strip().lower() not in {"", "0", "false", "#ref!", "#n/a", "n/a"}


def discontinued_value(value: Any) -> bool:
    value = display_value(value)
    if value is True:
        return True
    if value is None or value is False:
        return False
    return str(value).strip().lower() in {"1", "true", "yes", "y", "x"}


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("xlsx", type=Path)
    parser.add_argument("output", type=Path)
    args = parser.parse_args()

    workbook = load_workbook(args.xlsx, read_only=True, data_only=True)
    missing = set(SHEETS) - set(workbook.sheetnames)
    if missing:
        raise SystemExit(f"Missing required sheets: {', '.join(sorted(missing))}")

    counts: dict[str, int] = {}
    with args.output.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.writer(handle, dialect="excel-tab", quoting=csv.QUOTE_MINIMAL, lineterminator="\n")
        for sheet_name, config in SHEETS.items():
            worksheet = workbook[sheet_name]
            count = 0
            for row_number, cells in enumerate(worksheet.iter_rows(min_row=3), start=3):
                by_letter = {get_column_letter(cell.column): display_value(cell.value) for cell in cells if cell.value is not None}
                if not any(is_business_value(by_letter.get(letter)) for letter in config["business_columns"]):
                    continue

                row_data = {letter: value for letter, value in by_letter.items() if value is not None}
                for letter, legacy_key in config["legacy"].items():
                    if letter in row_data:
                        row_data[legacy_key] = row_data[letter]

                typed = {field: (by_letter.get(letter) if letter else None) for field, letter in config["typed"].items()}
                typed["discontinued"] = discontinued_value(typed["discontinued"])
                values = [
                    WORKBOOK_ID, sheet_name, row_number, config["tracker_type"],
                    *[typed.get(field) for field in TYPED_FIELDS],
                    json.dumps(row_data, ensure_ascii=False, separators=(",", ":")),
                ]
                writer.writerow(["\\N" if value is None else ("true" if value is True else "false" if value is False else value) for value in values])
                count += 1
            counts[sheet_name] = count

    print(json.dumps({"output": str(args.output), "counts": counts, "total": sum(counts.values())}))


if __name__ == "__main__":
    main()
